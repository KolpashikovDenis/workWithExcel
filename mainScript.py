from openpyxl import Workbook, load_workbook
from matplotlib import pyplot as plt
from matplotlib import ticker
import os
import numpy as np

def space_to_tab(line):
    c = []
    i = 0
    while i < len(line):
        if line[i] == ' ':
            c.append('\t')
            while line[i] == ' ':
                i += 1
            c.append(line[i])
        else:
            c.append(line[i])
        i += 1

    line = ''.join(c)
    return line

def show_graphs(t, cpu, qcpu, p_memused, p_swpused, p_avgqu_sz, p_await, p_svctm, p_util, p_net_rx, p_net_tx, p_avg1, p_avg5, p_avg15):
    np_x = np.asarray(t)
    X = len(t) - len(t) // 7

    #CPU
    cpu_y1 = np.asarray(cpu)
    qcpu_y2 = np.asarray(qcpu)

    #LoadAverage
    avg1_y = np.asarray(p_avg1)
    avg5_y = np.asarray(p_avg5)
    avg15_y= np.asarray(p_avg15)

    # mem and SWop USED
    memsused_y = np.asarray(p_memused)
    swpused_y = np.asarray(p_swpused)

    # очереди дисковой подсистемы
    avgqu_sz_y = np.asarray(p_avgqu_sz)

    # Среднее время чтения/записи
    await_y = np.asarray(p_await)
    svctm_y = np.asarray(p_svctm)

    util_y = np.asarray(p_util)

    # Утилизация сетевых интерфейсов
    net_rx_y = np.asarray(p_net_rx)
    net_tx_y = np.asarray(p_net_tx)

    fig=plt.figure(figsize=(10, 7))

    # Утилизация CPU
    ax00 = fig.add_subplot(3, 2, 1)
    ax00.plot(np_x, cpu_y1, color='blue', label='Утилизация CPU')
    ax00.plot(np_x, qcpu_y2, color='green', label='Очередь CPU')
    ax00.axvline(x=np_x[-1], color = 'red')
    ax00.set_xlabel('Продолжительность теста, ч:мм')
    ax00.set_ylabel('Утилизация CPU, %')
    ax00.set_title('Утилизация CPU')
    ax00.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax00.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    ylim = ax00.get_ylim()
    sub_ax00 = ax00.twinx()
    sub_ax00.plot([], [])
    sub_ax00.set_ylabel('Очереди CPU, шт')
    sub_ax00.set_ylim(ylim[0], ylim[1])
    sub_ax00.tick_params(labelsize=6, labelrotation=90)
    plt.grid()

    # Утилизация памяти и файла подкачки
    ax01=fig.add_subplot(3, 2, 2)
    ax01.plot(np_x, memsused_y, color='blue', label='Утилизация памяти')
    ax01.plot([], [], color='green', label='Утилизация подкачки')
    ax01.axvline(x=np_x[-1], color='red')
    ax01.set_xlabel('Продолжительность теста, ч:мм')
    ax01.set_ylabel('Утилизация памяти, %')
    ax01.set_title('Утилизация памяти')
    ax01.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax01.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    sub_ax01 = ax01.twinx()
    sub_ax01.plot(np_x, swpused_y)
    sub_ax01.set_ylabel('Утилизация подкачки, %')
    sub_ax01.xaxis.set_major_locator(ticker.MultipleLocator(5))
    sub_ax01.tick_params(labelsize=6, labelrotation=90)

    plt.grid()

    # очередь дисковой подсистемы
    ax10=fig.add_subplot(3, 2, 3)
    ax10.plot(np_x, avgqu_sz_y, color='blue', label='Очередь дисковой подсистемы')
    ax10.axvline(x=np_x[-1], color='red')
    ax10.set_xlabel('Продолжительность теста, ч:мм')
    ax10.set_ylabel('Очередь дисковой\nподсистемы, шт')
    ax10.set_title('Очередь дисковой подсистемы')
    ax10.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax10.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    plt.grid()

    # Load Average
    ax11=fig.add_subplot(3, 2, 4)
    ax11.plot(np_x, avg1_y, color='blue', label='за 1 минуту')
    ax11.plot(np_x, avg5_y, color='green', label='за 5 минут')
    ax11.plot(np_x, avg15_y, color='purple', label='за 15 минут')
    ax11.axvline(x=np_x[-1], color='red')
    ax11.set_xlabel('Продолжительность теста, ч:мм')
    ax11.set_ylabel('Значение коэффициента')
    ax11.set_title('Load Average')
    ax11.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax11.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    plt.grid()

    # Среднее время чтения и записи
    ax20=fig.add_subplot(3, 2, 5)
    ax20.plot(np_x, await_y, color='blue', label='среднее время выполнения чтения/записи')
    ax20.plot(np_x, svctm_y, color='green', label='среднее время обслуживания чтения/записи')
    ax20.axvline(x=np_x[-1], color='red')
    ax20.set_xlabel('Продолжительность теста, ч:мм')
    ax20.set_ylabel('Среднее время\nчтения/записи')
    ax20.set_title('Среднее время чтения/записи')
    ax20.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax20.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    plt.grid()

    # Утилизация сетевого интерфейса
    ax21=fig.add_subplot(3, 2, 6)
    ax21.plot(np_x, net_rx_y, color='blue', label='Получаемые данные')
    ax21.plot(np_x, net_tx_y, color='green', label='Передаваемые данные')
    ax21.axvline(x=np_x[-1], color='red')
    ax21.set_xlabel('Продолжительность теста, ч:мм')
    ax21.set_ylabel('Передаваемые данные')
    ax21.set_title('Утилизация сетевого интерфейса')
    ax21.xaxis.set_major_locator(ticker.MultipleLocator(5))
    ax21.tick_params(labelsize=6, labelrotation=90)
    plt.legend(loc='upper left')
    plt.grid()

    plt.subplots_adjust(wspace=0.3, hspace=0.5)
    plt.show()


# =========   MAIN CODE ===============================
# Формируем имя файла
# f_name = 'Report_'+datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M-%S")+'.xlsx'
f_name_report = "Report.xlsx"
f_name_with_data = "sar_mpgu_izh.csv"

# Вспомогательный список
data = []

# Данные для графика утилизации CPU
cpu = []
q_cpu = []

# Списки для хранения данных по использованию оперативки и свопа
memsused = []
swpused = []

# Списки в которых хранятся устредненные данные по использованию дисков
l_avgqu_sz = []
l_await = []
l_svctm = []
l_util = []

# В этом списке содержатся усредненные данные по использованию сетевых интерфейсов
net_rx = []
net_tx = []

# список для Load Average
ld_avg_1 = []
ld_avg_5 = []
ld_avg_15 = []

# список для хранения времени
t = []

# делаем екселевский файл
Items = ['Graphs', 'CPU', 'MEM', 'DISK', 'NET', 'LOAD_AVG']
#
# if os.path.exists(f_name_report):
#     wb_report = load_workbook(f_name_report)
#     for i in range(1, len(Items)):
#         tmp_sheet = wb_report[Items[i]]
#         wb_report.remove(tmp_sheet)
# else:
#     wb_report = Workbook()
#     for item in Items:
#         wb_report.create_sheet(item)
#     wdel = wb_report['Sheet']
#     wb_report.remove(wdel)
wb_report = Workbook()
for item in Items:
    wb_report.create_sheet(item)
wdel = wb_report['Sheet']
wb_report.remove(wdel)

line = str('')

# Данные для CPU
with open(f_name_with_data, 'r') as infile:
    # active_sheet = wb_report[Items[1]]
    # Утилизация CPU, idle в %
    while True:
        line = infile.readline().strip()
        if '%idle' in line:
            # data = space_to_tab(line).split('\t')
            # del data[1:10]
            # data[0] = data[0][:5]
            # active_sheet.cell(row=1, column=1).value = data[0]
            # active_sheet.cell(row=1, column=2).value = data[1]
            break

    r = 2
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        if 'all' in line:
            data = space_to_tab(line).split('\t')
            del data[1:10]
            data[0] = data[0][:5]
            # active_sheet.cell(row=r, column=1).value = data[0]
            t.append(data[0])
            a = 100.0 - float(data[1].replace(',', '.'))
            # active_sheet.cell(row=r, column=2).value = a
            cpu.append(a)
            r += 1

    # очереди CPU
    while True:
        line = infile.readline().strip()
        if 'runq-sz' in line:
            # data = space_to_tab(line).split('\t')
            # del data[2:]
            # active_sheet.cell(row=1, column=3).value = data[1]
            break

    r = 2
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[2:]
        a = float(data[1].replace(',', '.'))
        # active_sheet.cell(row=r, column=3).value = a
        q_cpu.append(a)
        r += 1

# Данные для МЕМ
with open(f_name_with_data, 'r') as infile:
    # active_sheet = wb_report[Items[2]]
    while True:
        line = infile.readline().strip()
        if 'memused' in line:
            # data = space_to_tab(line).split('\t')
            # del data[1:3]
            # del data[2:]
            # data[0] = data[0][:5]
            # active_sheet.cell(row=1, column=1).value = data[0]
            # active_sheet.cell(row=1, column=2).value = data[1]
            break

    r = 2
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[1:3]
        del data[2:]
        data[0] = data[0][:5]
        # active_sheet.cell(row=r, column=1).value = data[0]
        a = float(data[1].replace(',', '.'))
        # active_sheet.cell(row=r, column=2).value = a
        memsused.append(a)
        r += 1

    while True:
        line = infile.readline().strip()
        if 'swpused' in line:
            data = space_to_tab(line).split('\t')
            del data[1:3]
            del data[2:]
            # active_sheet.cell(row=1, column=3).value = data[1]
            break

    r = 2
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[1:3]
        del data[2:]
        a = float(data[1].replace(',', '.'))
        # active_sheet.cell(row=r, column=3).value = a
        swpused.append(a)
        r += 1

# Данные для среднего времени чтения/записи
with open(f_name_with_data, 'r') as infile:
    active_sheet = wb_report[Items[3]]
    while True:
        line = infile.readline().strip()
        if 'DEV' in line:
            # data = space_to_tab(line).split('\t')
            # del data[1:6]
            # data[0] = data[0][:5]
            # active_sheet.cell(row=1, column=1).value = data[0]
            # active_sheet.cell(row=1, column=2).value = data[1]
            # active_sheet.cell(row=1, column=3).value = data[2]
            # active_sheet.cell(row=1, column=4).value = data[3]
            # active_sheet.cell(row=1, column=5).value = data[4]
            break

    # САМЫЙ ЖОПОШНЫЙ УЧАСТОК КОДА
    map = {}
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[2:6]
        data[0] = data[0][:5]
        key = data[0]
        del data[0:2]
        if key in map.keys():
            map[key].append(data)
        else:
            map[key] = [data]

    size = len(map['09:21'])
    avg_map = {}
    rowNum = 2
    for key in map.keys():
        _avgqu_sz = 0.0
        _await = 0.0
        _svctm = 0.0
        _util = 0.0
        for i in map[key]:
            _avgqu_sz += float(i[0].replace(',', '.'))
            _await += float(i[1].replace(',', '.'))
            _svctm += float(i[2].replace(',', '.'))
            _util += float(i[3].replace(',', '.'))

        l_avgqu_sz.append(_avgqu_sz / size)
        l_await.append(_await / size)
        l_svctm.append(_svctm / size)
        l_util.append(_util / size)

        avg_map[key] = [_avgqu_sz / size, _await / size, _svctm / size, _util / size]
        # active_sheet.cell(row=rowNum, column=1).value = key
        # active_sheet.cell(row=rowNum, column=2).value = avg_map[key][0]
        # active_sheet.cell(row=rowNum, column=3).value = avg_map[key][1]
        # active_sheet.cell(row=rowNum, column=4).value = avg_map[key][2]
        # active_sheet.cell(row=rowNum, column=5).value = avg_map[key][3]
        rowNum += 1

# Усредненные данные по сетевым интерфейсам
with open(f_name_with_data, 'r') as infile:
    active_sheet = wb_report[Items[4]]
    while True:
        line = infile.readline().strip()
        if 'IFACE' in line:
            # data = space_to_tab(line).split('\t')
            # del data[1:4]
            # del data[3:]
            # data[0] = data[0][:5]
            # active_sheet.cell(row=1, column=1).value = data[0]
            # active_sheet.cell(row=1, column=2).value = data[1]
            # active_sheet.cell(row=1, column=3).value = data[2]
            break

    map = {}
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[1:4]
        del data[3:]
        data[0] = data[0][:5]
        key = data[0]
        del data[0]
        if key in map.keys():
            map[key].append(data)
        else:
            map[key] = [data]

    size = len(map['09:21'])
    avg_map = {}
    rowNum = 2
    for key in map.keys():
        rxkB = 0.0
        txkB = 0.0
        for i in map[key]:
            rxkB += float(i[0].replace(',', '.'))
            txkB += float(i[1].replace(',', '.'))

        net_rx.append(rxkB / size)
        net_tx.append(txkB / size)

        avg_map[key] = [rxkB / size, txkB / size]
        # active_sheet.cell(row=rowNum, column=1).value = key
        # active_sheet.cell(row=rowNum, column=2).value = avg_map[key][0]
        # active_sheet.cell(row=rowNum, column=3).value = avg_map[key][1]
        rowNum += 1

# Динамика Load Average
with open(f_name_with_data, 'r') as infile:
    active_sheet = wb_report[Items[5]]
    while True:
        line = infile.readline().strip()
        if 'runq-sz' in line:
            # data = space_to_tab(line).split('\t')
            # del data[1:3]
            # data[0] = data[0][:5]
            # active_sheet.cell(row=1, column=1).value = data[0]
            # active_sheet.cell(row=1, column=2).value = data[1]
            # active_sheet.cell(row=1, column=3).value = data[2]
            # active_sheet.cell(row=1, column=4).value = data[3]
            break

    r = 2
    while True:
        line = infile.readline().strip()
        if 'Average' in line:
            break
        data = space_to_tab(line).split('\t')
        del data[1:3]
        data[0] = data[0][:5]

        # active_sheet.cell(row=r, column=1).value = data[0]
        # active_sheet.cell(row=r, column=2).value = float(data[1].replace(',', '.'))
        # active_sheet.cell(row=r, column=3).value = float(data[2].replace(',', '.'))
        # active_sheet.cell(row=r, column=4).value = float(data[3].replace(',', '.'))
        ld_avg_1.append(float(data[1].replace(',', '.')))
        ld_avg_5.append(float(data[2].replace(',', '.')))
        ld_avg_15.append(float(data[3].replace(',', '.')))
        r += 1

# Рисуем все графики оптом!
show_graphs(t, cpu, q_cpu, memsused, swpused, l_avgqu_sz, l_await, l_svctm, l_util, net_rx, net_tx, ld_avg_1, ld_avg_5, ld_avg_15)

# wb_report.save(f_name_report)